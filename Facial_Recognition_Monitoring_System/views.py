from django.conf import settings
from django.shortcuts import render, redirect
from django.db import connection
from django.contrib import messages
import openpyxl
from django.core.files.storage import FileSystemStorage
from datetime import datetime
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.views.decorators.cache import never_cache
from django.core.mail import send_mail
import os
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

def home(request):
    return render(request, "index.html")

@never_cache
def admin_login(request):
    if request.method == 'POST':
        try:
            user_id1 = request.POST.get('username')
            password1 = request.POST.get('password')
            user = authenticate(request, username=user_id1, password=password1)
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT user_id, user_name, password FROM account_loginadmin WHERE user_id = %s",
                    [user_id1]
                )
                row = cursor.fetchone()
                if user is not None:
                    login(request, user)
                    request.session['user_id'] = row[0]
                    request.session['username'] = row[1]
                    return redirect('/admin-dashboard/')
                else:
                    return render(request, "admin_login.html", {"error": "Invalid data.."})
        except Exception as e:
            return render(request, "admin_login.html", {"error": "Data fetch error.."})

    else:
        return render(request, "admin_login.html")


@never_cache
def teacher_login(request):
    if request.method == 'POST':
        try:
            user_id1 = request.POST.get('username')
            password1 = request.POST.get('password')
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT user_id, user_name, password FROM account_loginteacher WHERE user_id = %s",
                    [user_id1]
                )
                row = cursor.fetchone()
                if row and row[2] == password1:
                    request.session['user_id'] = row[0]
                    request.session['username'] = row[1]
                    return redirect('/teacher-dashboard/')
                else:
                    return render(request, "teacher_login.html", {"error": "Invalid data.."})
        except Exception as e:
            return render(request, "teacher_login.html", {"error": "Data fetch error.."})

    else:
        return render(request, "teacher_login.html")


@never_cache
def student_login(request):
    if request.method == 'POST':
        try:
            user_id1 = request.POST.get('username')
            password1 = request.POST.get('password')
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT user_id, user_name, password FROM account_loginstudent WHERE user_id = %s",
                    [user_id1]
                )
                row = cursor.fetchone()
                if row and row[2] == password1:
                    request.session['user_id'] = row[0]
                    request.session['username'] = row[1]
                    return redirect('/student-dashboard/')
                else:
                    return render(request, "student_login.html", {"error": "Invalid data.."})
        except Exception as e:
            return render(request, "student_login.html", {"error": "Data fetch error.."})

    else:
        return render(request, "student_login.html")

def _get_excel_row_values(row, expected_length):
    """
    Helper to safely extract values from an Excel row, replacing None with empty string,
    and padding/truncating to expected_length.
    """
    values = [cell if cell is not None else "" for cell in row]
    # Pad with empty strings if row is short
    if len(values) < expected_length:
        values += [""] * (expected_length - len(values))
    # Truncate if row is too long
    return values[:expected_length]

@never_cache
def admin_dashboard(request):
    if 'user_id' not in request.session:
        return redirect("index")
    username = request.session.get('username')
    request.session['user_id'] = request.session.get('user_id')
    message = datetime.now()
    if request.method == 'POST':
        if request.POST.get("action") == "add":
            table_name = "table_" + str(request.session.get('user_id'))
            try:
                with connection.cursor() as cursor:
                    cursor.execute(f'''
                        CREATE TABLE IF NOT EXISTS `{table_name}` (
                            id VARCHAR(20) PRIMARY KEY, 
                            name VARCHAR(25),
                            email VARCHAR(25),
                            department VARCHAR(25),
                            subject VARCHAR(25),
                            semester VARCHAR(25),
                            year VARCHAR(4),
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        );
                    ''')
                excel_file = request.FILES.get('excelFile')
                if excel_file:
                    wb = openpyxl.load_workbook(excel_file)
                    sheet = wb.active
                    with connection.cursor() as cursor:
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            # Use helper to handle empty cells and missing columns
                            student_id, name, email, department, subject, semester, year = _get_excel_row_values(row, 7)
                            if not student_id:
                                continue
                            cursor.execute(
                                f'''INSERT INTO `{table_name}` (id, name, email, department, subject, semester, year)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s);''',
                                [str(student_id), str(name), str(email), str(department), str(subject), str(semester), str(year)]
                            )
                            cursor.execute('''
                                INSERT IGNORE INTO account_loginteacher (user_id, user_name, password, email, department, year)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            ''', [
                                str(student_id),
                                str(name),
                                str(student_id),
                                str(email),
                                str(department),
                                str(year)
                            ])
                    message = "Excel data uploaded successfully."
                else:
                    student_id = request.POST.get('userId')
                    name = request.POST.get('name')
                    email = request.POST.get('email')
                    department = request.POST.get('department')
                    subject = request.POST.get('subject')
                    semester = request.POST.get('semester')
                    year = request.POST.get('year')
                    if student_id and name and email and department and subject and semester and year:
                        with connection.cursor() as cursor:
                            cursor.execute(
                                f'''INSERT INTO `{table_name}` (id, name, email, department, subject, semester, year)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s);''',
                                [student_id, name, email, department, subject, semester, year]
                            )
                            cursor.execute('''
                                INSERT IGNORE INTO account_loginteacher (user_id, user_name, password, email, department, year)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            ''', [
                                str(student_id),
                                str(name),
                                str(student_id),
                                str(email),
                                str(department),
                                str(year)
                            ])
                        message = "Manual data added successfully."
                    else:
                        message = "Invalid data"
            except Exception as e:
                message = "Duplicate entry or Error occurred."
        elif request.POST.get("action") == "pass":
            try:
                user_id = request.session.get('user_id')
                old_password = request.POST.get('old_password')
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')

                if new_password != confirm_password:
                    message = "Confirm Password does not match"
                else:
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT password FROM account_loginadmin WHERE user_id = %s", [user_id])
                        row = cursor.fetchone()

                        if row and row[0] == old_password:
                            user = request.user
                            user.set_password(new_password)
                            user.save()
                            update_session_auth_hash(request, user)
                            cursor.execute(
                                "UPDATE account_loginadmin SET password = %s WHERE user_id = %s",
                                [new_password, user_id]
                            )
                            message = "Password changed successfully"
                        else:
                            message = "Old password is incorrect"
            except Exception as e:
                message = "An error occurred while changing password"
    return render(request, "admin-dashboard.html", {"username": username, "message": message})

@never_cache
def teacher_dashboard(request):
    if 'user_id' not in request.session:
        return redirect("index")
    username = request.session.get('username')
    request.session['user_id'] = request.session.get('user_id')
    message = datetime.now()
    if request.method == 'POST':
        if request.POST.get("action") == "add":
            table_name = "table_" + str(request.session.get('user_id'))
            try:
                with connection.cursor() as cursor:
                    cursor.execute(f'''
                        CREATE TABLE IF NOT EXISTS `{table_name}` (
                            id VARCHAR(20) PRIMARY KEY, 
                            name VARCHAR(25),
                            email VARCHAR(25),
                            department VARCHAR(25),
                            subject VARCHAR(25),
                            semester VARCHAR(25),
                            year VARCHAR(4),
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        );
                    ''')
                excel_file = request.FILES.get('excelFile')
                if excel_file:
                    wb = openpyxl.load_workbook(excel_file)
                    sheet = wb.active
                    with connection.cursor() as cursor:
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            # Use helper to handle empty cells and missing columns
                            student_id, name, email, department, subject, semester, year = _get_excel_row_values(row, 7)
                            if not student_id:
                                continue
                            cursor.execute(
                                f'''INSERT INTO `{table_name}` (id, name, email, department, subject, semester, year)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s);''',
                                [str(student_id), str(name), str(email), str(department), str(subject), str(semester), str(year)]
                            )
                            cursor.execute('''
                                INSERT IGNORE INTO account_loginstudent (user_id, user_name, password, email, department, year)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            ''', [
                                str(student_id),
                                str(name),
                                str(student_id),
                                str(email),
                                str(department),
                                str(year)
                            ])
                    message = "Excel data uploaded successfully."
                else:
                    student_id = request.POST.get('userId')
                    name = request.POST.get('name')
                    email = request.POST.get('email')
                    department = request.POST.get('department')
                    subject = request.POST.get('subject')
                    semester = request.POST.get('semester')
                    year = request.POST.get('year')
                    if student_id and name and email and subject and semester and year:
                        with connection.cursor() as cursor:
                            cursor.execute(
                                f'''INSERT INTO `{table_name}` (id, name, email, department, subject, semester, year)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s);''',
                                [student_id, name, email, department, subject, semester, year]
                            )
                            cursor.execute('''
                                INSERT IGNORE INTO account_loginstudent (user_id, user_name, password, email, department, year)
                                VALUES (%s, %s, %s, %s, %s, %s)
                            ''', [
                                str(student_id),
                                str(name),
                                str(student_id),
                                str(email),
                                str(department),
                                str(year)
                            ])
                        message = "Manual data added successfully."
                    else:
                        message = "Invalid data"
            except Exception as e:
                message = "Duplicate entry or Error occurred."
        elif request.POST.get("action") == "pass":
            try:
                user_id = request.session.get('user_id')
                old_password = request.POST.get('old_password')
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')

                if new_password != confirm_password:
                    message = "Confirm Password does not match"
                else:
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT password FROM account_loginteacher WHERE user_id = %s", [user_id])
                        row = cursor.fetchone()

                        if row and row[0] == old_password:
                            cursor.execute(
                                "UPDATE account_loginteacher SET password = %s WHERE user_id = %s",
                                [new_password, user_id]
                            )
                            message = "Password changed successfully"
                        else:
                            message = "Old password is incorrect"
            except Exception as e:
                message = "An error occurred while changing password"
    return render(request, "teacher-dashboard.html", {"username": username, "message": message})

@never_cache
def student_dashboard(request):
    if 'user_id' not in request.session:
        return redirect("index")
    username = request.session.get('username')
    message = datetime.now()
    if request.method == 'POST':
        if request.POST.get("action") == "pass":
            try:
                user_id = request.session.get('user_id')
                old_password = request.POST.get('old_password')
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')

                if new_password != confirm_password:
                    message = "Confirm Password does not match"
                else:
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT password FROM account_loginstudent WHERE user_id = %s", [user_id])
                        row = cursor.fetchone()

                        if row and row[0] == old_password:
                            cursor.execute(
                                "UPDATE account_loginstudent SET password = %s WHERE user_id = %s",
                                [new_password, user_id]
                            )
                            message = "Password changed successfully"
                        else:
                            message = "Old password is incorrect"
            except Exception as e:
                message = "An error occurred while changing password"
                
        if request.POST.get("action") == "update_profile_pic":
            try:
                user_id = request.session.get('user_id')
                profile_pic = request.FILES.get('profile_pic')
                if profile_pic:
                    # Generate a unique filename
                    ext = profile_pic.name.split('.')[-1]
                    filename = f"{user_id}.{ext}"
                    file_path = os.path.join('profile_pics', filename)
                    # Delete previous profile picture if exists
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT profile_pic FROM account_loginstudent WHERE user_id = %s", [user_id])
                        prev_row = cursor.fetchone()
                        if prev_row and prev_row[0]:
                            prev_file_path = prev_row[0]
                            if default_storage.exists(prev_file_path):
                                default_storage.delete(prev_file_path)
                    # Save the new file
                    path = default_storage.save(file_path, ContentFile(profile_pic.read()))
                    # Update the database with the new profile picture path
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "UPDATE account_loginstudent SET profile_pic = %s WHERE user_id = %s",
                            [file_path, user_id]
                        )
                    message = "Profile picture updated successfully"
                else:
                    message = "No profile picture uploaded"
            except Exception as e:
                message = "An error occurred while updating profile picture"
    # Access the student's profile picture for display
    profile_pic_url = None
    try:
        user_id = request.session.get('user_id')
        with connection.cursor() as cursor:
            cursor.execute("SELECT profile_pic FROM account_loginstudent WHERE user_id = %s", [user_id])
            row = cursor.fetchone()
            if row and row[0]:
                profile_pic_url = row[0]
    except Exception:
        profile_pic_url = None
    return render(request, "student-dashboard.html", {"username": username, "message": message, "profile_pic_url": profile_pic_url})

def logout(request):
    if 'user_id' in request.session:
        del request.session['user_id']
    return redirect('index')

def teacher_forget(request):
    data = {
        "data": ""
    }
    if request.method == 'POST':
        if request.POST.get("action") == "id":
            try:
                user_id1 = request.POST.get('user_id')
                data["data"] = "Invalid id"
                with connection.cursor() as cursor:
                    cursor.execute(
                        "SELECT user_id, email, password FROM account_loginteacher WHERE user_id = %s",
                        [user_id1]
                    )
                    row = cursor.fetchone()
                    if row and len(row) > 0:
                        send_mail(
                            "PASSWORD",
                            "Password is : " + row[2],
                            "collegeproject0007@gmail.com",
                            [row[1]],
                            fail_silently=False,
                        )
                        data["data"] = row[1]
            except:
                pass
    return render(request, "forgetpass/teacher.html", data)

def student_forget(request):
    data = {
        "data": ""
    }
    if request.method == 'POST':
        if request.POST.get("action") == "id":
            try:
                user_id1 = request.POST.get('user_id')
                data["data"] = "Invalid id"
                with connection.cursor() as cursor:
                    cursor.execute(
                        "SELECT user_id, email, password FROM account_loginstudent WHERE user_id = %s",
                        [user_id1]
                    )
                    row = cursor.fetchone()
                    if row and len(row) > 0:
                        send_mail(
                            "PASSWORD",
                            "Password is : " + row[2],
                            "collegeproject0007@gmail.com",
                            [row[1]],
                            fail_silently=False,
                        )
                        data["data"] = row[1]
            except:
                pass
    return render(request, "forgetpass/student.html", data)

def admin_forget(request):
    data = {
        "data": ""
    }
    if request.method == 'POST':
        if request.POST.get("action") == "id":
            try:
                user_id1 = request.POST.get('user_id')
                data["data"] = "Invalid id"
                with connection.cursor() as cursor:
                    cursor.execute(
                        "SELECT user_id, email, password FROM account_loginadmin WHERE user_id = %s",
                        [user_id1]
                    )
                    row = cursor.fetchone()
                    if row and len(row) > 0:
                        send_mail(
                            "PASSWORD",
                            "Password is : " + row[2],
                            "collegeproject0007@gmail.com",
                            [row[1]],
                            fail_silently=False,
                        )
                        data["data"] = row[1]
            except:
                pass
    return render(request, "forgetpass/admin.html", data)


def update_teacher(request):
    data = {
        "data": '',
        "department":"department",
        "semester":"semester",
        "year":"year",
    }
    table = "table_" + str(request.session.get('user_id'))
    if request.method == 'POST':
        if request.POST.get("action") == "filter":
            try:
                # Validate table name to prevent SQL injection
                if not table.isidentifier():
                    data["all_teachers"] = []
                else:
                    subject = request.POST.get('subject')
                    department = request.POST.get('department')
                    semester = request.POST.get('semester')
                    year = request.POST.get('year')
                    with connection.cursor() as cursor:
                        query = f"""
                            SELECT id, name, email, department, subject, semester, year
                            FROM `{table}`
                            WHERE subject = %s AND department = %s AND semester = %s AND year = %s
                        """
                        cursor.execute(query, [subject, department, semester, year])
                        all_teachers = cursor.fetchall()
                        data["all_teachers"] = all_teachers
            except Exception:
                data["all_teachers"] = []
            data["department"] = request.POST.get('department')
            data["subject"] = request.POST.get('subject')
            data["semester"] = request.POST.get('semester')
            data["year"] = request.POST.get('year')
                
        if request.POST.get("action") == "id":
            try:
                teacher_id = request.POST.get('user_id')
                data["data"] = "Not Found"
                with connection.cursor() as cursor:
                    cursor.execute(
                        f"""
                        SELECT id, name, email, department, subject, semester, year
                        FROM {table}
                        WHERE id = %s
                        """,
                        [teacher_id],
                    )
                    row = cursor.fetchone()
                    if row and len(row) > 0:
                        data["id"] = row[0]
                        data["name"] = row[1]
                        data["email"] = row[2]
                        data["department"] = row[3]
                        data["subject"] = row[4]
                        data["semester"] = row[5]
                        data["year"] = row[6]
                        data["data"] = "Found"
            except:
                pass
        if request.POST.get("action") == "update":
            try:
                teacher_id = request.POST.get('userId')
                name = request.POST.get('name')
                email = request.POST.get('email')
                department = request.POST.get('department')
                subject = request.POST.get('subject')
                semester = request.POST.get('semester')
                year = request.POST.get('year')
                if teacher_id and name and email and department and subject and semester and year:
                    with connection.cursor() as cursor:
                        cursor.execute(
                            f'''UPDATE `{table}` 
                                SET name=%s, email=%s, department=%s, subject=%s, semester=%s, year=%s
                                WHERE id=%s;''',
                            [name, email, department, subject, semester, year, teacher_id]
                        )
                    data["data"] = "Update Successful"
                else:
                    data["data"] = "Invalid data for update"
            except:
                pass

        if request.POST.get("action") == "delete":
            try:
                teacher_id = request.POST.get('userId')
                if teacher_id:
                    with connection.cursor() as cursor:
                        cursor.execute(
                            f"DELETE FROM `{table}` WHERE id = %s", [teacher_id]
                        )
                    data["data"] = "Delete Successful"
                else:
                    data["data"] = "No ID provided for deletion"
            except:
                pass
    return render(request, 'edit_details/teacher.html', data)

def update_student(request):
    data = {
        "data": '',
        "department":"department",
        "semester":"semester",
        "year":"year",
    }
    table = "table_" + str(request.session.get('user_id'))
    if request.method == 'POST':
        if request.POST.get("action") == "filter":
            try:
                # Validate table name to prevent SQL injection
                if not table.isidentifier():
                    data["all_teachers"] = []
                else:
                    subject = request.POST.get('subject')
                    department = request.POST.get('department')
                    semester = request.POST.get('semester')
                    year = request.POST.get('year')
                    with connection.cursor() as cursor:
                        query = f"""
                            SELECT id, name, email, department, subject, semester, year
                            FROM `{table}`
                            WHERE subject = %s AND department = %s AND semester = %s AND year = %s
                        """
                        cursor.execute(query, [subject, department, semester, year])
                        all_students = cursor.fetchall()
                        data["all_students"] = all_students
            except Exception:
                data["all_students"] = []
            data["department"] = request.POST.get('department')
            data["subject"] = request.POST.get('subject')
            data["semester"] = request.POST.get('semester')
            data["year"] = request.POST.get('year')
        if request.POST.get("action") == "id":
            try:
                student_id = request.POST.get('user_id')
                data["data"] = "Not Found"
                with connection.cursor() as cursor:
                    cursor.execute(
                        f"""
                        SELECT id, name, email, department, subject, semester, year
                        FROM `{table}`
                        WHERE id = %s
                        """,
                        [student_id],
                    )
                    row = cursor.fetchone()
                    if row and len(row) > 0:
                        data["id"] = row[0]
                        data["name"] = row[1]
                        data["email"] = row[2]
                        data["department"] = row[3]
                        data["subject"] = row[4]
                        data["semester"] = row[5]
                        data["year"] = row[6]
                        data["data"] = "Found"
            except:
                pass
        if request.POST.get("action") == "update":
            try:
                student_id = request.POST.get('userId')
                name = request.POST.get('name')
                email = request.POST.get('email')
                department = request.POST.get('department')
                subject = request.POST.get('subject')
                semester = request.POST.get('semester')
                year = request.POST.get('year')
                if student_id and name and email and department and subject and semester and year:
                    with connection.cursor() as cursor:
                        cursor.execute(
                            f'''UPDATE `{table}` 
                                SET name=%s, email=%s, department=%s, subject=%s, semester=%s, year=%s
                                WHERE id=%s;''',
                            [name, email, department, subject, semester, year, student_id]
                        )
                    data["data"] = "Update Successful"
                else:
                    data["data"] = "Invalid data for update"
            except:
                pass

        if request.POST.get("action") == "delete":
            try:
                student_id = request.POST.get('userId')
                if student_id:
                    with connection.cursor() as cursor:
                        cursor.execute(
                            f"DELETE FROM `{table}` WHERE id = %s", [student_id]
                        )
                    data["data"] = "Delete Successful"
                else:
                    data["data"] = "No ID provided for deletion"
            except:
                pass
    print(data["data"])
    return render(request, 'edit_details/student.html', data)


